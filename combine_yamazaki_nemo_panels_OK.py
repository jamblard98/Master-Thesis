
"""
Script d'éxécution : 
python3  ~/Thesis/Codes/combine_yamazaki_nemo_panels_OK.py \
  --yamazaki-csv-dir ~/Thesis/Yamazaki/Datas/Datas_OK \
  --nemo-csv-dir ~/Thesis/NEMO/Datas/Datas_OK \
  --out-dir ~/Thesis/Outputs \
  --season DJF \
  --depth-min 10 \
  --depth-max 200 \
  --hist-y0 1900 \
  --hist-y1 1970

Panneaux : Yamazaki (gauche) vs NEMO (droite)
Profils d'anomalie par mer.

LOGIQUE LONGITUDE 0-360° (sens horaire) :
- Ross: 160° → 210° (50°)
- Weddell: 300° → 340° (40°)
- Bellingshausen-Amundsen: 230° → 300° (70°)
- Davis: 80° → 100° (20°)
"""

import argparse
import glob
from pathlib import Path
import numpy as np
import pandas as pd

# IMPORTANT: backend non-interactif pour éviter les erreurs Qt/Wayland en environnement headless
import matplotlib
matplotlib.use("Agg")

import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from matplotlib.lines import Line2D
from matplotlib.transforms import Bbox

# === AJOUTS (Excel + styles) ===
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


SEASON_MONTHS = {
    'DJF': [12, 1, 2],
    'MAM': [3, 4, 5],
    'JJA': [6, 7, 8],
    'SON': [9, 10, 11],
    'ALL': list(range(1, 13)),
}

# =============================================================================
# AJOUT : BANDES DE LATITUDE (pour profils latitudinaux)
# =============================================================================

LAT_BANDS = [
    ("60-70°S", -70.0, -60.0),  # [-70, -60[
    ("70-80°S", -80.0, -70.0),  # [-80, -70[
]

# =============================================================================
# DÉFINITION DES MERS (0-360° sens horaire)
# =============================================================================

SEAS = {
    'Ross': {
        'lat_min': -78, 'lat_max': -70,
        'lon_start': 160,   # 160°E
        'lon_end': 210,     # 210° = -150°W
        'label': 'Ross'
    },
    'Weddell': {
        'lat_min': -78, 'lat_max': -60,
        'lon_start': 300,   # 300° = -60°W
        'lon_end': 340,     # 340° = -20°W
        'label': 'Weddell'
    },
    'Bellingshausen-Amundsen': {
        'lat_min': -75, 'lat_max': -65,
        'lon_start': 230,   # 230° = -130°W
        'lon_end': 300,     # 300° = -60°W
        'label': 'Bellingshausen/\nAmundsen'
    },
    'Davis': {
        'lat_min': -68, 'lat_max': -65,
        'lon_start': 80,    # 80°E
        'lon_end': 100,     # 100°E
        'label': 'Davis'
    }
}

SEA_ORDER = ['Bellingshausen-Amundsen', 'Weddell', 'Davis', 'Ross']

DECADES = [
    (1900, 1910), (1910, 1920), (1920, 1930), (1930, 1940),
    (1940, 1950), (1950, 1960), (1960, 1970)
]

DECADE_COLORS = ['#1f77b4', '#ff7f0e', '#2ca02c',
                 '#d62728', '#9467bd', '#8c564b', '#e377c2']

# =============================================================================
# FONCTIONS UTILITAIRES
# =============================================================================

def _safe_read_csvs(pattern: str) -> pd.DataFrame:
    files = sorted(glob.glob(pattern))

    print(f" Pattern: {pattern}")
    print(f" Fichiers trouvés: {len(files)}")

    dfs = []
    total_lines = 0

    for f in files:
        try:
            df = pd.read_csv(f)
            if df.dropna(how='all').empty:
                continue
            dfs.append(df)
            fname = Path(f).name
            print(f"    {fname}: {len(df):,} lignes")
            total_lines += len(df)
        except Exception as e:
            print(f"  Lecture échouée: {f} ({e})")

    print(f"TOTAL chargé: {total_lines:,} lignes depuis {len(dfs)} fichiers")

    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def normalize_longitude_360(lon):
    """Convertit longitude [-180,180] vers [0,360]."""
    return lon % 360


def filter_by_sea(df, sea_config):
    """
    Filtre DataFrame par mer avec logique 0-360°.

    Args:
        df: DataFrame avec 'hist_lat' et 'hist_lon' (en [-180,180])
        sea_config: Dict avec lat_min, lat_max, lon_start, lon_end

    Returns:
        DataFrame filtré
    """
    # Convertir longitudes en 0-360
    df = df.copy()
    df['lon_360'] = normalize_longitude_360(df['hist_lon'])

    # Filtrer latitude
    df_sea = df[
        (df['hist_lat'] >= sea_config['lat_min']) &
        (df['hist_lat'] < sea_config['lat_max'])
    ]

    # Filtrer longitude (sens horaire)
    lon_start = sea_config['lon_start']
    lon_end = sea_config['lon_end']

    if lon_start <= lon_end:
        # Arc normal
        df_sea = df_sea[
            (df_sea['lon_360'] >= lon_start) &
            (df_sea['lon_360'] < lon_end)
        ]
    else:
        # Passage par 0°
        df_sea = df_sea[
            (df_sea['lon_360'] >= lon_start) |
            (df_sea['lon_360'] < lon_end)
        ]

    return df_sea


# =============================================================================
# FONCTIONS DE TRACÉ
# =============================================================================

def _plot_yamazaki_sea_on_ax(ax, df_sea, sea_name, season):
    """Dessine profils d'anomalie Yamazaki pour une mer."""
    has_data = False

    for idx, (dec0, dec1) in enumerate(DECADES):
        df_dec = df_sea[(df_sea['hist_year'] >= dec0) &
                        (df_sea['hist_year'] < dec1)]
        if len(df_dec) == 0:
            continue

        has_data = True

        zmin = max(10, int(np.floor(df_dec['hist_depth_m'].min())))
        zmax = min(200, int(np.ceil(df_dec['hist_depth_m'].max())))
        depth_bins = np.arange(zmin, zmax + 10, 10)
        if len(depth_bins) < 2:
            continue
        depth_centers = (depth_bins[:-1] + depth_bins[1:]) / 2

        anom_mean = []
        for i in range(len(depth_bins) - 1):
            m = ((df_dec['hist_depth_m'] >= depth_bins[i]) &
                 (df_dec['hist_depth_m'] < depth_bins[i + 1]))
            sub = df_dec[m]
            if len(sub) > 0:
                anom_mean.append(
                    (sub['hist_temperature'] - sub['yamazaki_T']).mean()
                )
            else:
                anom_mean.append(np.nan)

        anom_mean = np.array(anom_mean)
        valid = ~np.isnan(anom_mean)

        if valid.any():
            ax.plot(anom_mean[valid], depth_centers[valid], '-',
                    color=DECADE_COLORS[idx], linewidth=2,
                    marker='o', markersize=4, alpha=0.9)

    if not has_data:
        ax.text(0.5, 0.5, "Pas de données", transform=ax.transAxes,
                ha='center', va='center', fontsize=9, color='gray')

    ax.set_xlim(-2.5, 2.5)
    ax.set_ylim(200, 10)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.axvline(0, color='black', linewidth=0.8, alpha=0.4)


def _plot_nemo_sea_on_ax(ax, df_sea, sea_name, season):
    """Dessine profils d'anomalie NEMO pour une mer."""
    has_data = False

    for idx, (dec0, dec1) in enumerate(DECADES):
        df_dec = df_sea[(df_sea['hist_year'] >= dec0) &
                        (df_sea['hist_year'] < dec1)]
        if len(df_dec) == 0:
            continue

        # Grouper et moyenner années récentes
        group_cols = ['hist_year', 'hist_month', 'hist_day',
                      'hist_lat', 'hist_lon', 'hist_depth_m', 'nemo_hist_T']

        df_grouped = df_dec.groupby(group_cols, as_index=False).agg({
            'nemo_recent_T': 'mean'
        })
        df_grouped.rename(columns={'nemo_recent_T': 'nemo_recent_mean'},
                          inplace=True)
        df_grouped['delta_T'] = (
            df_grouped['nemo_hist_T'] - df_grouped['nemo_recent_mean']
        )

        has_data = True

        zmin = max(10, int(np.floor(df_grouped['hist_depth_m'].min())))
        zmax = min(200, int(np.ceil(df_grouped['hist_depth_m'].max())))
        depth_bins = np.arange(zmin, zmax + 10, 10)
        if len(depth_bins) < 2:
            continue
        depth_centers = (depth_bins[:-1] + depth_bins[1:]) / 2

        anom_mean = []
        for i in range(len(depth_bins) - 1):
            m = ((df_grouped['hist_depth_m'] >= depth_bins[i]) &
                 (df_grouped['hist_depth_m'] < depth_bins[i + 1]))
            sub = df_grouped[m]
            if len(sub) > 0:
                anom_mean.append(sub['delta_T'].mean())
            else:
                anom_mean.append(np.nan)

        anom_mean = np.array(anom_mean)
        valid = ~np.isnan(anom_mean)

        if valid.any():
            ax.plot(anom_mean[valid], depth_centers[valid], '-',
                    color=DECADE_COLORS[idx], linewidth=2,
                    marker='o', markersize=4, alpha=0.9)

    if not has_data:
        ax.text(0.5, 0.5, "Pas de données", transform=ax.transAxes,
                ha='center', va='center', fontsize=9, color='gray')

    ax.set_xlim(-2.5, 2.5)
    ax.set_ylim(200, 10)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.axvline(0, color='black', linewidth=0.8, alpha=0.4)


# =============================================================================
# PROFILS MOYENS & STATISTIQUES VERTICALES (TABLEAU)  [conservé mais optionnel]
# =============================================================================

def _compute_yamazaki_profile(df_sea, year0, year1,
                              depth_min=10, depth_max=200,
                              bin_size=10):
    """Profil moyen ΔT_obs(z) = obs - Yamazaki sur [year0, year1)."""
    df_dec = df_sea[
        (df_sea['hist_year'] >= year0) &
        (df_sea['hist_year'] < year1)
    ].copy()
    if df_dec.empty:
        return np.array([]), np.array([])

    df_dec = df_dec[
        (df_dec['hist_depth_m'] >= depth_min) &
        (df_dec['hist_depth_m'] <= depth_max)
    ]
    if df_dec.empty:
        return np.array([]), np.array([])

    zmin = max(depth_min, int(np.floor(df_dec['hist_depth_m'].min())))
    zmax = min(depth_max, int(np.ceil(df_dec['hist_depth_m'].max())))
    if zmax <= zmin:
        return np.array([]), np.array([])

    depth_bins = np.arange(zmin, zmax + bin_size, bin_size)
    if len(depth_bins) < 2:
        return np.array([]), np.array([])

    depth_centers = (depth_bins[:-1] + depth_bins[1:]) / 2
    anom_mean = []

    for i in range(len(depth_bins) - 1):
        m = ((df_dec['hist_depth_m'] >= depth_bins[i]) &
             (df_dec['hist_depth_m'] < depth_bins[i + 1]))
        sub = df_dec[m]
        if len(sub) > 0:
            vals = sub['hist_temperature'] - sub['yamazaki_T']
            anom_mean.append(vals.mean())
        else:
            anom_mean.append(np.nan)

    return depth_centers, np.array(anom_mean)


def _compute_nemo_profile(df_sea, year0, year1,
                          depth_min=10, depth_max=200,
                          bin_size=10):
    """Profil moyen ΔT_NEMO(z) = NEMO_hist - NEMO_recent_mean sur [year0, year1)."""
    df_dec = df_sea[
        (df_sea['hist_year'] >= year0) &
        (df_sea['hist_year'] < year1)
    ].copy()
    if df_dec.empty:
        return np.array([]), np.array([])

    group_cols = ['hist_year', 'hist_month', 'hist_day',
                  'hist_lat', 'hist_lon', 'hist_depth_m', 'nemo_hist_T']
    df_grouped = df_dec.groupby(group_cols, as_index=False).agg({
        'nemo_recent_T': 'mean'
    })
    df_grouped.rename(columns={'nemo_recent_T': 'nemo_recent_mean'},
                      inplace=True)
    df_grouped['delta_T'] = (
        df_grouped['nemo_hist_T'] - df_grouped['nemo_recent_mean']
    )

    df_grouped = df_grouped[
        (df_grouped['hist_depth_m'] >= depth_min) &
        (df_grouped['hist_depth_m'] <= depth_max)
    ]
    if df_grouped.empty:
        return np.array([]), np.array([])

    zmin = max(depth_min, int(np.floor(df_grouped['hist_depth_m'].min())))
    zmax = min(depth_max, int(np.ceil(df_grouped['hist_depth_m'].max())))
    if zmax <= zmin:
        return np.array([]), np.array([])

    depth_bins = np.arange(zmin, zmax + bin_size, bin_size)
    if len(depth_bins) < 2:
        return np.array([]), np.array([])

    depth_centers = (depth_bins[:-1] + depth_bins[1:]) / 2
    anom_mean = []

    for i in range(len(depth_bins) - 1):
        m = ((df_grouped['hist_depth_m'] >= depth_bins[i]) &
             (df_grouped['hist_depth_m'] < depth_bins[i + 1]))
        sub = df_grouped[m]
        if len(sub) > 0:
            anom_mean.append(sub['delta_T'].mean())
        else:
            anom_mean.append(np.nan)

    return depth_centers, np.array(anom_mean)


def _align_profiles(depths1, vals1, depths2, vals2):
    """Aligne deux profils sur les profondeurs communes et vire les NaN."""
    if len(depths1) == 0 or len(depths2) == 0:
        return np.array([]), np.array([])

    df1 = pd.DataFrame({'depth': depths1, 'v1': vals1})
    df2 = pd.DataFrame({'depth': depths2, 'v2': vals2})
    dfm = df1.merge(df2, on='depth', how='inner')
    if dfm.empty:
        return np.array([]), np.array([])

    mask = np.isfinite(dfm['v1']) & np.isfinite(dfm['v2'])
    dfm = dfm[mask]
    if dfm.empty:
        return np.array([]), np.array([])

    return dfm['v1'].to_numpy(), dfm['v2'].to_numpy()


def _compute_vertical_stats_for_sea(sea_name,
                                    df_sea_yam,
                                    df_sea_nem,
                                    depth_min=10,
                                    depth_max=200):
    """
    Calcule les stats verticales pour une mer :
    - r_1900_1929, r_1930_1969
    - RMSE_1900_1969
    - amplitude_obs, amplitude_nemo
    - N_obs, N_mod
    """
    d_obs_1_z, d_obs_1_v = _compute_yamazaki_profile(
        df_sea_yam, 1900, 1930, depth_min, depth_max
    )
    d_mod_1_z, d_mod_1_v = _compute_nemo_profile(
        df_sea_nem, 1900, 1930, depth_min, depth_max
    )
    v1_obs, v1_mod = _align_profiles(d_obs_1_z, d_obs_1_v,
                                     d_mod_1_z, d_mod_1_v)
    if len(v1_obs) > 0:
        r_1900_1929 = float(np.corrcoef(v1_obs, v1_mod)[0, 1])
    else:
        r_1900_1929 = np.nan

    d_obs_2_z, d_obs_2_v = _compute_yamazaki_profile(
        df_sea_yam, 1930, 1970, depth_min, depth_max
    )
    d_mod_2_z, d_mod_2_v = _compute_nemo_profile(
        df_sea_nem, 1930, 1970, depth_min, depth_max
    )
    v2_obs, v2_mod = _align_profiles(d_obs_2_z, d_obs_2_v,
                                     d_mod_2_z, d_mod_2_v)
    if len(v2_obs) > 0:
        r_1930_1969 = float(np.corrcoef(v2_obs, v2_mod)[0, 1])
    else:
        r_1930_1969 = np.nan

    d_obs_all_z, d_obs_all_v = _compute_yamazaki_profile(
        df_sea_yam, 1900, 1970, depth_min, depth_max
    )
    d_mod_all_z, d_mod_all_v = _compute_nemo_profile(
        df_sea_nem, 1900, 1970, depth_min, depth_max
    )

    N_obs = int(np.isfinite(d_obs_all_v).sum()) if len(d_obs_all_v) > 0 else 0
    N_mod = int(np.isfinite(d_mod_all_v).sum()) if len(d_mod_all_v) > 0 else 0

    v_all_obs, v_all_mod = _align_profiles(d_obs_all_z, d_obs_all_v,
                                           d_mod_all_z, d_mod_all_v)
    if len(v_all_obs) > 0:
        diff = v_all_obs - v_all_mod
        RMSE_1900_1969 = float(np.sqrt(np.mean(diff ** 2)))
    else:
        RMSE_1900_1969 = np.nan

    if N_obs > 0:
        amplitude_obs = float(np.nanmax(d_obs_all_v) - np.nanmin(d_obs_all_v))
    else:
        amplitude_obs = np.nan

    if N_mod > 0:
        amplitude_nemo = float(np.nanmax(d_mod_all_v) - np.nanmin(d_mod_all_v))
    else:
        amplitude_nemo = np.nan

    sea_label = SEAS[sea_name]['label'].replace('\n', ' ')

    return {
        'Sea': sea_label,
        'r_1900_1929': r_1900_1929,
        'r_1930_1969': r_1930_1969,
        'RMSE_1900_1969': RMSE_1900_1969,
        'amplitude_obs': amplitude_obs,
        'amplitude_nemo': amplitude_nemo,
        'N_obs': N_obs,
        'N_mod': N_mod
    }


def _save_vertical_stats_table(stats_rows, out_dir: Path, silent=False):
    """Sauvegarde le tableau de stats verticales en CSV + print terminal."""
    if not stats_rows:
        if not silent:
            print("\n  Aucune statistique verticale à sauvegarder.")
        return

    df_stats = pd.DataFrame(stats_rows)

    cols = [
        'Sea',
        'r_1900_1929',
        'r_1930_1969',
        'RMSE_1900_1969',
        'amplitude_obs',
        'amplitude_nemo',
        'N_obs',
        'N_mod'
    ]
    df_stats = df_stats[cols]

    out_csv = out_dir / "NEMO_Yamazaki_comparison.csv"
    df_stats.to_csv(out_csv, index=False)

    if not silent:
        print("\n" + "="*72)
        print("TABLEAU DE COMPARAISON NEMO vs YAMAZAKI (statistiques verticales)")
        print("="*72)
        pd.options.display.float_format = "{:,.3f}".format
        print(df_stats.to_string(index=False))
        print("="*72 + "\n")
        print(f"Tableau comparatif sauvegardé : {out_csv.name}")


# =============================================================================
# AJOUTS : TABLEAUX STATS (Tableau 1 / M / D) + Excel + PNG propre
# =============================================================================

def _fmt_depth_label(d0, d1):
    # Conserver exactement les labels + ajout shallow (10–50, 50–100)
    if abs(d0 - 10) < 1e-9 and abs(d1 - 50) < 1e-9:
        return "10–50 m"
    if abs(d0 - 50) < 1e-9 and abs(d1 - 100) < 1e-9:
        return "50–100 m"
    if abs(d0 - 10) < 1e-9 and abs(d1 - 100) < 1e-9:
        return "10–100 m"
    if abs(d0 - 100) < 1e-9 and abs(d1 - 200) < 1e-9:
        return "100–200 m"
    if abs(d0 - 10) < 1e-9 and abs(d1 - 200) < 1e-9:
        return "10–200 m"
    return f"{int(d0)}–{int(d1)} m"


def _compute_obs_delta_points(df_obs: pd.DataFrame) -> pd.DataFrame:
    """
    OBS: ΔT = T_histo - T_récente = hist_temperature - yamazaki_T
    Ici on garde le point tel quel (pas de regroupement).
    """
    d = df_obs.copy()
    d['T_histo'] = d['hist_temperature']
    d['T_recent'] = d['yamazaki_T']
    d['delta_T'] = d['T_histo'] - d['T_recent']
    return d


def _compute_nemo_delta_points(df_nem: pd.DataFrame) -> pd.DataFrame:
    """
    NEMO: ΔT = T_histo - T_récente = nemo_hist_T - mean(nemo_recent_T)
    Regroupement par "observation historique unique" (mêmes group_cols que profils NEMO).
    """
    d = df_nem.copy()
    group_cols = ['hist_year', 'hist_month', 'hist_day',
                  'hist_lat', 'hist_lon', 'hist_depth_m', 'nemo_hist_T']
    g = d.groupby(group_cols, as_index=False).agg({'nemo_recent_T': 'mean'})
    g.rename(columns={'nemo_recent_T': 'nemo_recent_mean'}, inplace=True)
    g['T_histo'] = g['nemo_hist_T']
    g['T_recent'] = g['nemo_recent_mean']
    g['delta_T'] = g['T_histo'] - g['T_recent']
    return g


def _summ_stats_scatter(delta: pd.Series):
    """
    Stats demandées (sans médiane/iqr) :
    - ⟨ΔT⟩
    - σ(ΔT)
    - % ΔT < 0  (au-dessus de y=x selon votre convention de signe via ΔT)
    - % ΔT > 0  (en-dessous)
    - N
    """
    x = delta.to_numpy(dtype=float)
    x = x[np.isfinite(x)]
    n = int(x.size)
    if n == 0:
        return (np.nan, np.nan, np.nan, np.nan, 0)
    mean = float(np.mean(x))
    std = float(np.std(x, ddof=1)) if n >= 2 else np.nan
    pct_neg = float(100.0 * np.mean(x < 0.0))
    pct_pos = float(100.0 * np.mean(x > 0.0))
    return (mean, std, pct_neg, pct_pos, n)


def _build_table1(dataset_name: str, df_points: pd.DataFrame) -> pd.DataFrame:
    """
    Tableau 1 : Mer (1ère colonne) + Période fusionnée + Profondeur (3 niveaux),
    colonnes stats: ⟨ΔT⟩, σ(ΔT), %ΔT<0, N.
    """
    periods = [(1900, 1930, "1900–1929"),
               (1930, 1970, "1930–1969"),
               (1900, 1970, "1900–1969")]
    depths = [(10, 100), (100, 200), (10, 200)]

    rows = []
    for sea_name in SEA_ORDER:
        sea_label = SEAS[sea_name]['label'].replace('\n', ' ')
        sea_conf = SEAS[sea_name]

        df_sea = filter_by_sea(df_points, sea_conf)

        for (y0, y1, plab) in periods:
            df_p = df_sea[(df_sea['hist_year'] >= y0) & (df_sea['hist_year'] < y1)]
            for (d0, d1) in depths:
                df_pd = df_p[(df_p['hist_depth_m'] >= d0) & (df_p['hist_depth_m'] <= d1)]
                mean, std, pct_neg, _, n = _summ_stats_scatter(df_pd['delta_T']) if len(df_pd) else (np.nan, np.nan, np.nan, np.nan, 0)
                rows.append({
                    'Mer': sea_label,
                    'Période': plab,
                    'Profondeur': _fmt_depth_label(d0, d1),
                    '⟨ΔT⟩': mean,
                    'σ(ΔT)': std,
                    '% ΔT < 0': pct_neg,
                    'N': n
                })

    return pd.DataFrame(rows)


def _build_tableM(dataset_name: str, df_points: pd.DataFrame) -> pd.DataFrame:
    """
    Tableau M (annexe) — Statistiques sectorielles complètes
    Dimensions : Mer, Période, Profondeur
    Colonnes : ⟨ΔT⟩, σ(ΔT), % ΔT < 0, N
    """
    return _build_table1(dataset_name, df_points)


def _build_tableD(dataset_name: str, df_points: pd.DataFrame) -> pd.DataFrame:
    """
    Tableau D (annexe) — Évolution décennale par mer
    Choix : une seule tranche 10–200 m
    Lignes : Décennies
    Colonnes : ⟨ΔT⟩, σ(ΔT), N
    """
    depth0, depth1 = 10, 200
    rows = []
    for sea_name in SEA_ORDER:
        sea_label = SEAS[sea_name]['label'].replace('\n', ' ')
        sea_conf = SEAS[sea_name]
        df_sea = filter_by_sea(df_points, sea_conf)
        df_sea = df_sea[(df_sea['hist_depth_m'] >= depth0) & (df_sea['hist_depth_m'] <= depth1)]

        for (dec0, dec1) in DECADES:
            dec_lab = f"{dec0}–{dec1-1}"
            sub = df_sea[(df_sea['hist_year'] >= dec0) & (df_sea['hist_year'] < dec1)]
            mean, std, _, _, n = _summ_stats_scatter(sub['delta_T']) if len(sub) else (np.nan, np.nan, np.nan, np.nan, 0)
            rows.append({
                'Mer': sea_label,
                'Décennie': dec_lab,
                '⟨ΔT⟩': mean,
                'σ(ΔT)': std,
                'N': n
            })

    return pd.DataFrame(rows)


# =============================================================================
# AJOUT : TABLEAU "LAT×DECADE" (statistiques par décennie et bande de latitude)
# =============================================================================

def _build_table_latband_decade(dataset_name: str, df_points: pd.DataFrame) -> pd.DataFrame:
    """
    Stats par décennie et bande de latitude (figure jointe) :
    - ⟨ΔT⟩ (anomalie moyenne)
    - σ(ΔT)
    - N

    Remarque : les filtres saison/profondeur/période ont déjà été appliqués en amont
    (df_points construit à partir de df_yam_filtered / df_nem_filtered).
    """
    rows = []
    for (band_label, lat_min, lat_max) in LAT_BANDS:
        sub_band = df_points[(df_points['hist_lat'] >= lat_min) & (df_points['hist_lat'] < lat_max)]
        for (dec0, dec1) in DECADES:
            dec_lab = f"{dec0}–{dec1-1}"
            sub = sub_band[(sub_band['hist_year'] >= dec0) & (sub_band['hist_year'] < dec1)]
            mean, std, _, _, n = _summ_stats_scatter(sub['delta_T']) if len(sub) else (np.nan, np.nan, np.nan, np.nan, 0)
            rows.append({
                'Bande de latitude': band_label,
                'Décennie': dec_lab,
                '⟨ΔT⟩': mean,
                'σ(ΔT)': std,
                'N': n
            })
    return pd.DataFrame(rows)


# =============================================================================
# AJOUT : TABLEAU "LAT×PERIODE×PROFONDEUR"
# =============================================================================

def _build_table_latband_period_depth(dataset_name: str, df_points: pd.DataFrame) -> pd.DataFrame:
    """
    Résultats par bande de latitude POUR DES PERIODES (1900–1929, 1930–1969, 1900–1969)
    et POUR DES PROFONDEURS (10–100, 100–200, 10–200).

    Colonnes :
    - Bande de latitude
    - Période
    - Profondeur
    - N
    - ⟨ΔT⟩
    - σ(ΔT)

    Remarque : df_points est déjà filtré en saison et sur [depth_min, depth_max] + [y0,y1)
    en amont ; ici on sous-sélectionne à nouveau par période et tranches de profondeur.
    """
    periods = [
        (1900, 1930, "1900–1929"),
        (1930, 1970, "1930–1969"),
        (1900, 1970, "1900–1969"),
    ]
    depths = [(10, 100), (100, 200), (10, 200)]

    rows = []
    for (band_label, lat_min, lat_max) in LAT_BANDS:
        sub_band = df_points[(df_points['hist_lat'] >= lat_min) & (df_points['hist_lat'] < lat_max)].copy()

        for (y0, y1, plab) in periods:
            sub_p = sub_band[(sub_band['hist_year'] >= y0) & (sub_band['hist_year'] < y1)]

            for (d0, d1) in depths:
                sub_pd = sub_p[(sub_p['hist_depth_m'] >= d0) & (sub_p['hist_depth_m'] <= d1)]
                if len(sub_pd) == 0:
                    mean, std, n = (np.nan, np.nan, 0)
                else:
                    x = pd.to_numeric(sub_pd['delta_T'], errors='coerce')
                    x = x[np.isfinite(x)]
                    n = int(x.size)
                    if n == 0:
                        mean, std = (np.nan, np.nan)
                    else:
                        mean = float(np.mean(x))
                        std = float(np.std(x, ddof=1)) if n >= 2 else np.nan

                rows.append({
                    'Bande de latitude': band_label,
                    'Période': plab,
                    'Profondeur': _fmt_depth_label(d0, d1),
                    'N': n,
                    '⟨ΔT⟩': mean,
                    'σ(ΔT)': std,
                })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(['Bande de latitude', 'Période', 'Profondeur']).reset_index(drop=True)


# =============================================================================
# AJOUTS : TABLEAUX "SCATTER SO" (pas de mers) — OBS, NEMO, et combiné
# =============================================================================

def _build_scatter_SO_table(dataset_name: str, df_points: pd.DataFrame, depths=None) -> pd.DataFrame:
    """
    Tableau SCATTER (Ocean Austral, sans distinction de mer) :
    Périodes : 1900–1929, 1930–1969, 1900–1969

    Profondeurs par défaut : 10–100, 100–200, 10–200
    Profondeurs custom possibles via depths=[(10,50),(50,100)] etc.

    Stats :
    - ⟨ΔT⟩  (ΔT = T_histo - T_récente, point par point)
    - σ(ΔT)
    - % ΔT < 0  (historique plus froid => réchauffement)
    - % ΔT > 0  (historique plus chaud => refroidissement)
    - N (paires)
    """
    periods = [(1900, 1930, "1900–1929"),
               (1930, 1970, "1930–1969"),
               (1900, 1970, "1900–1969")]

    if depths is None:
        depths = [(10, 100), (100, 200), (10, 200)]

    rows = []
    for (y0, y1, plab) in periods:
        df_p = df_points[(df_points['hist_year'] >= y0) & (df_points['hist_year'] < y1)]
        for (d0, d1) in depths:
            df_pd = df_p[(df_p['hist_depth_m'] >= d0) & (df_p['hist_depth_m'] <= d1)]
            mean, std, pct_neg, pct_pos, n = _summ_stats_scatter(df_pd['delta_T']) if len(df_pd) else (np.nan, np.nan, np.nan, np.nan, 0)
            rows.append({
                'Dataset': dataset_name,
                'Période': plab,
                'Profondeur': _fmt_depth_label(d0, d1),
                '⟨ΔT⟩': mean,
                'σ(ΔT)': std,
                '% ΔT < 0': pct_neg,
                '% ΔT > 0': pct_pos,
                'N': n
            })
    return pd.DataFrame(rows)


def _apply_sheet_style(ws: Worksheet, header_row=1):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # largeurs colonnes raisonnables
    col_widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            col_widths[cell.column_letter] = max(col_widths.get(cell.column_letter, 8), min(45, len(val) + 2))
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # hauteur lignes
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18


def _merge_cells_for_table1(ws: Worksheet, sea_col=1, period_col=2, start_row=2):
    """
    Fusionne :
    - Mer : blocs de 9 lignes (3 périodes × 3 profondeurs)
    - Période : blocs de 3 lignes (3 profondeurs)
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        sea_val = ws.cell(row=r, column=sea_col).value
        if sea_val is None:
            r += 1
            continue

        # fin du bloc Mer
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=sea_col).value == sea_val:
            r2 += 1
        sea_end = r2 - 1

        if sea_end > r:
            ws.merge_cells(start_row=r, start_column=sea_col, end_row=sea_end, end_column=sea_col)

        # fusion des périodes à l'intérieur du bloc Mer
        rp = r
        while rp <= sea_end:
            per_val = ws.cell(row=rp, column=period_col).value
            if per_val is None:
                rp += 1
                continue
            rp2 = rp
            while rp2 <= sea_end and ws.cell(row=rp2, column=period_col).value == per_val:
                rp2 += 1
            per_end = rp2 - 1
            if per_end > rp:
                ws.merge_cells(start_row=rp, start_column=period_col, end_row=per_end, end_column=period_col)
            rp = rp2

        r = sea_end + 1


def _merge_cells_for_tableD(ws: Worksheet, sea_col=1, start_row=2):
    """
    Fusionne la colonne 'Mer' dans TableD (blocs par mer sur les décennies).
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        sea_val = ws.cell(row=r, column=sea_col).value
        if sea_val is None:
            r += 1
            continue
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=sea_col).value == sea_val:
            r2 += 1
        end = r2 - 1
        if end > r:
            ws.merge_cells(start_row=r, start_column=sea_col, end_row=end, end_column=sea_col)
        r = end + 1


def _merge_cells_for_scatter_period(ws: Worksheet, period_col=1, start_row=2):
    """
    Fusionne la colonne 'Période' (blocs de N lignes de profondeur) dans Scatter_SO_*.
    (Fonction générique : fonctionne aussi bien pour 3 profondeurs que pour 2.)
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        v = ws.cell(row=r, column=period_col).value
        if v is None:
            r += 1
            continue
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=period_col).value == v:
            r2 += 1
        end = r2 - 1
        if end > r:
            ws.merge_cells(start_row=r, start_column=period_col, end_row=end, end_column=period_col)
        r = end + 1


def _merge_cells_for_scatter_combined(ws: Worksheet, dataset_col=1, period_col=2, start_row=2):
    """
    Scatter_SO_Combined :
    - fusion verticale de Dataset (OBS puis NEMO)
    - ET fusion de Période à l'intérieur de chaque bloc Dataset
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        dv = ws.cell(row=r, column=dataset_col).value
        if dv is None:
            r += 1
            continue

        # bloc Dataset
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=dataset_col).value == dv:
            r2 += 1
        ds_end = r2 - 1
        if ds_end > r:
            ws.merge_cells(start_row=r, start_column=dataset_col, end_row=ds_end, end_column=dataset_col)

        # fusion Période dans le bloc Dataset
        rp = r
        while rp <= ds_end:
            pv = ws.cell(row=rp, column=period_col).value
            if pv is None:
                rp += 1
                continue
            rp2 = rp
            while rp2 <= ds_end and ws.cell(row=rp2, column=period_col).value == pv:
                rp2 += 1
            pe = rp2 - 1
            if pe > rp:
                ws.merge_cells(start_row=rp, start_column=period_col, end_row=pe, end_column=period_col)
            rp = rp2

        r = ds_end + 1


# =============================================================================
# AJOUT : Fusion Excel pour "LAT×DECADE"
# =============================================================================

def _merge_cells_for_latband_decade(ws: Worksheet, band_col=1, start_row=2):
    """
    Fusionne la colonne 'Bande de latitude' dans le tableau LAT×DECADE.
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        v = ws.cell(row=r, column=band_col).value
        if v is None:
            r += 1
            continue
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=band_col).value == v:
            r2 += 1
        end = r2 - 1
        if end > r:
            ws.merge_cells(start_row=r, start_column=band_col, end_row=end, end_column=band_col)
        r = end + 1


# =============================================================================
# AJOUT : Fusion Excel pour "LAT×PERIODE×PROFONDEUR"
# =============================================================================

def _merge_cells_for_latband_period_depth(ws: Worksheet, band_col=1, period_col=2, start_row=2):
    """
    Fusionne :
    - Bande de latitude : blocs de 9 lignes (3 périodes × 3 profondeurs)
    - Période : blocs de 3 lignes (3 profondeurs) à l'intérieur de chaque bande
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        band_val = ws.cell(row=r, column=band_col).value
        if band_val is None:
            r += 1
            continue

        # fin du bloc bande
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=band_col).value == band_val:
            r2 += 1
        band_end = r2 - 1

        if band_end > r:
            ws.merge_cells(start_row=r, start_column=band_col, end_row=band_end, end_column=band_col)

        # fusion des périodes à l'intérieur du bloc bande
        rp = r
        while rp <= band_end:
            per_val = ws.cell(row=rp, column=period_col).value
            if per_val is None:
                rp += 1
                continue
            rp2 = rp
            while rp2 <= band_end and ws.cell(row=rp2, column=period_col).value == per_val:
                rp2 += 1
            per_end = rp2 - 1
            if per_end > rp:
                ws.merge_cells(start_row=rp, start_column=period_col, end_row=per_end, end_column=period_col)
            rp = rp2

        r = band_end + 1


def _write_df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    _apply_sheet_style(ws, header_row=1)
    return ws


# =============================================================================
# PNG: fusion visuelle (texte + bordures internes supprimées) + format 2 décimales fixes
# =============================================================================

def _cell_set_visible_edges(cell, edges: str):
    # Compatibilité Matplotlib (selon version)
    if hasattr(cell, "set_visible_edges"):
        cell.set_visible_edges(edges)
    else:
        cell.visible_edges = edges


def _cell_get_visible_edges(cell) -> str:
    if hasattr(cell, "get_visible_edges"):
        return cell.get_visible_edges()
    if hasattr(cell, "visible_edges"):
        return cell.visible_edges
    return "LRBT"


def _fmt_fixed(x, decimals=2):
    """Format fixe avec 'decimals' décimales (ex: 0.00)."""
    try:
        if x is None:
            return ""
        x = float(x)
        if not np.isfinite(x):
            return ""
        return f"{x:.{decimals}f}"
    except Exception:
        return ""


def _apply_visual_merges_matplotlib(table, df_raw: pd.DataFrame, visual_merge):
    """
    Applique une "fusion visuelle" sur un tableau matplotlib:
    - efface le texte répété (sauf première ligne du run),
    - supprime les bordures internes horizontales (T/B) dans la colonne concernée.
    visual_merge = list of dict:
      { "col": "<colname>", "groupby": ["colA","colB", ...] }
    """
    if not visual_merge:
        return

    df0 = df_raw.copy()

    col_index = {c: i for i, c in enumerate(df0.columns.tolist())}

    for spec in visual_merge:
        c = spec.get("col")
        groupby = spec.get("groupby", []) or []
        if c not in col_index:
            continue
        ci = col_index[c]

        # itère par groupes
        if groupby:
            groups = df0.groupby(groupby, sort=False)
            group_iters = [(k, g.index.to_list()) for k, g in groups]
        else:
            group_iters = [(None, df0.index.to_list())]

        for _, idxs in group_iters:
            if not idxs:
                continue
            # parcours séquentiel sur l'ordre actuel du DF
            run_start = idxs[0]
            prev = df0.loc[idxs[0], c]
            for ii in idxs[1:] + [None]:
                if ii is None:
                    run_end = idxs[-1]
                    same = True
                else:
                    v = df0.loc[ii, c]
                    same = (v == prev)

                if not same:
                    run_end = idxs[idxs.index(ii) - 1]

                if (ii is None) or (not same):
                    # appliquer fusion sur run [run_start, run_end] si longueur > 1
                    if run_end != run_start:
                        pos_map = {idx: p for p, idx in enumerate(df0.index.to_list())}
                        p0 = pos_map[run_start]
                        p1 = pos_map[run_end]
                        for p in range(p0, p1 + 1):
                            tr = p + 1
                            tc = ci
                            cell = table[(tr, tc)]
                            # effacer texte sauf première ligne
                            if p != p0:
                                cell.get_text().set_text("")
                            # supprimer bordures internes T/B
                            edges = set(_cell_get_visible_edges(cell))
                            edges.add("L"); edges.add("R")
                            if p != p0:
                                edges.discard("T")
                            if p != p1:
                                edges.discard("B")
                            _cell_set_visible_edges(cell, "".join([e for e in "LRBT" if e in edges]))

                    # reset run
                    if ii is not None:
                        run_start = ii
                        prev = df0.loc[ii, c]


def _save_table_png_generic(df: pd.DataFrame, out_png: Path, font_size=10, col_widths=None, visual_merge=None, decimals=2):
    """
    PNG sans titre, CROP EXACT au cadre du tableau (pas de blanc inutile),
    + fusion visuelle optionnelle via visual_merge.
    Format numérique : 'decimals' décimales fixes (par défaut 2).
    """
    PAD_PX = 0

    d = df.copy()

    # format numérique
    for c in ['⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'z']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce').map(lambda x: _fmt_fixed(x, decimals=decimals))
    if 'N' in d.columns:
        d['N'] = pd.to_numeric(d['N'], errors='coerce').map(lambda x: "" if not np.isfinite(x) else f"{int(x)}")

    fig_h = max(4.5, 0.38 * (len(d) + 1))
    fig_w = 14

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    ax.set_position([0, 0, 1, 1])

    table = ax.table(
        cellText=d.values.tolist(),
        colLabels=d.columns.tolist(),
        cellLoc='center',
        colLoc='center',
        loc='upper left'
    )
    table.auto_set_font_size(False)
    table.set_fontsize(font_size)

    for (r, c), cell in table.get_celld().items():
        cell.set_linewidth(0.8)
        if r == 0:
            cell.set_facecolor('#F2F2F2')
            cell.set_text_props(weight='bold')
            cell.set_height(0.06)
        else:
            cell.set_height(0.045)

    if col_widths is not None:
        for c in range(len(col_widths)):
            for r in range(len(d) + 1):
                table[(r, c)].set_width(col_widths[c])

    _apply_visual_merges_matplotlib(table, df_raw=df, visual_merge=visual_merge)

    fig.canvas.draw()
    renderer = fig.canvas.get_renderer()
    bbox = table.get_window_extent(renderer=renderer)

    if PAD_PX and PAD_PX > 0:
        bbox = Bbox.from_extents(
            bbox.x0 - PAD_PX, bbox.y0 - PAD_PX,
            bbox.x1 + PAD_PX, bbox.y1 + PAD_PX
        )

    bbox_inches = bbox.transformed(fig.dpi_scale_trans.inverted())

    fig.savefig(out_png, dpi=300, bbox_inches=bbox_inches, pad_inches=0, facecolor='white')
    plt.close(fig)


def _save_table1_like_png(df_table: pd.DataFrame, out_png: Path, font_size=10, decimals=2):
    """
    PNG Table1 / TableM : fusion visuelle Mer + Période (bords internes supprimés).
    """
    df2 = df_table[['Mer', 'Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', 'N']].copy()

    _save_table_png_generic(
        df2,
        out_png,
        font_size=font_size,
        col_widths=[0.18, 0.16, 0.14, 0.12, 0.12, 0.12, 0.08],
        visual_merge=[
            {"col": "Mer", "groupby": []},
            {"col": "Période", "groupby": ["Mer"]},
        ],
        decimals=decimals
    )


def _save_tableD_png(df_tableD: pd.DataFrame, out_png: Path, font_size=10, decimals=2):
    """
    PNG TableD : fusion visuelle Mer (blocs par mer).
    """
    df2 = df_tableD[['Mer', 'Décennie', '⟨ΔT⟩', 'σ(ΔT)', 'N']].copy()
    _save_table_png_generic(
        df2,
        out_png,
        font_size=font_size,
        col_widths=[0.22, 0.16, 0.14, 0.14, 0.08],
        visual_merge=[
            {"col": "Mer", "groupby": []},
        ],
        decimals=decimals
    )


# =============================================================================
# AJOUT : PNG LAT×DECADE (version agrégée uniquement)
# =============================================================================

def _save_table_latband_decade_png(df_latdec: pd.DataFrame, out_png: Path, font_size=10, decimals=2):
    df2 = df_latdec[['Bande de latitude', 'Décennie', '⟨ΔT⟩', 'σ(ΔT)', 'N']].copy()
    _save_table_png_generic(
        df2,
        out_png,
        font_size=font_size,
        col_widths=[0.22, 0.18, 0.14, 0.14, 0.08],
        visual_merge=[
            {"col": "Bande de latitude", "groupby": []},
        ],
        decimals=decimals
    )


# =============================================================================
# AJOUT : PNG LAT×PERIODE×PROFONDEUR
# =============================================================================

def _save_table_latband_period_depth_png(df_latpd: pd.DataFrame, out_png: Path, font_size=10, decimals=2):
    df2 = df_latpd[['Bande de latitude', 'Période', 'Profondeur', 'N', '⟨ΔT⟩', 'σ(ΔT)']].copy()
    _save_table_png_generic(
        df2,
        out_png,
        font_size=font_size,
        col_widths=[0.22, 0.18, 0.14, 0.08, 0.14, 0.14],
        visual_merge=[
            {"col": "Bande de latitude", "groupby": []},
            {"col": "Période", "groupby": ["Bande de latitude"]},
        ],
        decimals=decimals
    )


# =============================================================================
# AJOUTS : "CASE TABLES" (3 tableaux ciblés) + z-score leave-one-year-out pooled
# =============================================================================

def _loo_zscore_pooled(df: pd.DataFrame, year: int, delta_col='delta_T', year_col='hist_year'):
    """
    z-score leave-one-year-out (pooled sur les POINTS des autres années)
    z = (mean_year - mean_others) / std_others
    - mean_year : moyenne des points ΔT de l'année 'year'
    - mean_others/std_others : stats sur les points ΔT de toutes les autres années
    """
    sub_y = df[df[year_col] == year][delta_col].to_numpy(dtype=float)
    sub_y = sub_y[np.isfinite(sub_y)]
    if sub_y.size == 0:
        return np.nan

    others = df[df[year_col] != year][delta_col].to_numpy(dtype=float)
    others = others[np.isfinite(others)]
    if others.size < 2:
        return np.nan

    mean_y = float(np.mean(sub_y))
    mean_o = float(np.mean(others))
    std_o = float(np.std(others, ddof=1)) if others.size >= 2 else np.nan
    if (not np.isfinite(std_o)) or std_o <= 0:
        return np.nan
    return float((mean_y - mean_o) / std_o)


def _build_case_table_depth_years_with_z(
    obs_pts: pd.DataFrame,
    sea_name: str,
    y0: int,
    y1_inclusive: int,
    depth_bands,
):
    """
    Construit un tableau "Profondeur / Année / N / ⟨ΔT⟩ / σ(ΔT) / z"
    - z calculé avec leave-one-year-out pooled sur les points des autres années (résout Ross).
    depth_bands = list of tuples: (d0, d1, label_str)
    """
    sea_conf = SEAS[sea_name]
    df = filter_by_sea(obs_pts, sea_conf)
    df = df[(df['hist_year'] >= y0) & (df['hist_year'] <= y1_inclusive)].copy()

    rows = []
    for (d0, d1, dlab) in depth_bands:
        sub = df[(df['hist_depth_m'] >= d0) & (df['hist_depth_m'] <= d1)].copy()
        if sub.empty:
            continue

        years = sorted(sub['hist_year'].dropna().unique().tolist())
        for yy in years:
            sy = sub[sub['hist_year'] == yy]['delta_T'].to_numpy(dtype=float)
            sy = sy[np.isfinite(sy)]
            if sy.size == 0:
                continue
            n = int(sy.size)
            m = float(np.mean(sy))
            sd = float(np.std(sy, ddof=1)) if sy.size >= 2 else np.nan
            z = _loo_zscore_pooled(sub, int(yy), delta_col='delta_T', year_col='hist_year')
            rows.append({
                'Profondeur': dlab,
                'Année': int(yy),
                'N': n,
                '⟨ΔT⟩': m,
                'σ(ΔT)': sd,
                'z': z
            })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values(['Profondeur', 'Année']).reset_index(drop=True)
    return out


def _build_case_table_years_simple(
    obs_pts: pd.DataFrame,
    sea_name: str,
    y0: int,
    y1_inclusive: int,
    depth_band=(10, 200),
):
    """
    Tableau simple "Année / N / ⟨ΔT⟩ / σ(ΔT)" (ex: Weddell 1937–1957).
    """
    sea_conf = SEAS[sea_name]
    df = filter_by_sea(obs_pts, sea_conf)
    df = df[(df['hist_year'] >= y0) & (df['hist_year'] <= y1_inclusive)].copy()
    d0, d1 = depth_band
    df = df[(df['hist_depth_m'] >= d0) & (df['hist_depth_m'] <= d1)].copy()

    rows = []
    for yy in sorted(df['hist_year'].dropna().unique().tolist()):
        sy = df[df['hist_year'] == yy]['delta_T'].to_numpy(dtype=float)
        sy = sy[np.isfinite(sy)]
        if sy.size == 0:
            continue
        rows.append({
            'Année': int(yy),
            'N': int(sy.size),
            '⟨ΔT⟩': float(np.mean(sy)),
            'σ(ΔT)': float(np.std(sy, ddof=1)) if sy.size >= 2 else np.nan
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values(['Année']).reset_index(drop=True)
    return out


def _merge_cells_for_case_depth(ws: Worksheet, depth_col=1, start_row=2):
    """
    Fusionne la colonne 'Profondeur' pour les case tables (blocs par profondeur).
    """
    max_row = ws.max_row
    r = start_row
    while r <= max_row:
        v = ws.cell(row=r, column=depth_col).value
        if v is None:
            r += 1
            continue
        r2 = r
        while r2 <= max_row and ws.cell(row=r2, column=depth_col).value == v:
            r2 += 1
        end = r2 - 1
        if end > r:
            ws.merge_cells(start_row=r, start_column=depth_col, end_row=end, end_column=depth_col)
        r = end + 1


def _save_case_depth_table_png(df_case: pd.DataFrame, out_png: Path, font_size=12, decimals=2):
    """
    PNG case table avec fusion visuelle de 'Profondeur'.
    """
    _save_table_png_generic(
        df_case,
        out_png,
        font_size=font_size,
        col_widths=[0.22, 0.14, 0.10, 0.14, 0.14, 0.10],
        visual_merge=[{"col": "Profondeur", "groupby": []}],
        decimals=decimals
    )


def _save_case_year_table_png(df_case: pd.DataFrame, out_png: Path, font_size=12, decimals=2):
    """
    PNG case table simple (Weddell).
    """
    _save_table_png_generic(
        df_case,
        out_png,
        font_size=font_size,
        col_widths=[0.18, 0.14, 0.18, 0.18],
        visual_merge=None,
        decimals=decimals
    )


# =============================================================================
# GÉNÉRATION TABLES (Excel + PNG)
# =============================================================================

def _generate_stats_outputs(df_yam_filtered: pd.DataFrame,
                            df_nem_filtered: pd.DataFrame,
                            out_dir: Path,
                            season: str,
                            depth_min: int,
                            depth_max: int,
                            y0: int,
                            y1: int,
                            make_png=True,
                            enable_scatter_tables=True):
    """
    Produit:
    - Excel: Table1/M/D pour OBS et NEMO (1 fichier)
      + une version brute de chaque feuille: suffixe "_brute"
    - PNG: Table1 + TableM + TableD pour OBS et NEMO (si make_png)
      + Scatter SO (OBS / NEMO / Combined) avec fusion visuelle
    - AJOUT: Scatter SO "shallow bands" 10–50 & 50–100 (2 lignes par période)
    - AJOUT: LAT×DECADE (par bande de latitude, par décennie) — OBS/NEMO + _brute
      + PNG pour la version agrégée uniquement
    - AJOUT: LAT×PERIODE×PROFONDEUR (par bande de latitude) — OBS/NEMO + _brute   
      + PNG pour la version agrégée uniquement                                   
    - AJOUT: 3 "Case tables" (BA, Weddell, Ross) + z-score pooled (résout Ross vide)
    - Format PNG : 2 décimales fixes pour toutes les valeurs numériques.
    """
    PNG_DECIMALS = 2

    obs_pts = _compute_obs_delta_points(df_yam_filtered)
    nemo_pts = _compute_nemo_delta_points(df_nem_filtered)

    # Tables par mer
    t1_obs = _build_table1("OBS", obs_pts)
    t1_nem = _build_table1("NEMO", nemo_pts)

    tm_obs = _build_tableM("OBS", obs_pts)
    tm_nem = _build_tableM("NEMO", nemo_pts)

    td_obs = _build_tableD("OBS", obs_pts)
    td_nem = _build_tableD("NEMO", nemo_pts)

    # =============================================================================
    # LAT×DECADE (OBS/NEMO)
    # =============================================================================
    latdec_obs = _build_table_latband_decade("OBS", obs_pts)
    latdec_nem = _build_table_latband_decade("NEMO", nemo_pts)

    # =============================================================================
    # LAT×PERIODE×PROFONDEUR (OBS/NEMO)
    # =============================================================================
    latpd_obs = _build_table_latband_period_depth("OBS", obs_pts)
    latpd_nem = _build_table_latband_period_depth("NEMO", nemo_pts)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws1 = _write_df_to_sheet(wb, "Table1_OBS", t1_obs)
    _merge_cells_for_table1(ws1, sea_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "Table1_OBS_brute", t1_obs)

    ws2 = _write_df_to_sheet(wb, "Table1_NEMO", t1_nem)
    _merge_cells_for_table1(ws2, sea_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "Table1_NEMO_brute", t1_nem)

    wsm1 = _write_df_to_sheet(wb, "TableM_OBS", tm_obs)
    _merge_cells_for_table1(wsm1, sea_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "TableM_OBS_brute", tm_obs)

    wsm2 = _write_df_to_sheet(wb, "TableM_NEMO", tm_nem)
    _merge_cells_for_table1(wsm2, sea_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "TableM_NEMO_brute", tm_nem)

    wsd1 = _write_df_to_sheet(wb, "TableD_OBS", td_obs)
    _merge_cells_for_tableD(wsd1, sea_col=1, start_row=2)
    _write_df_to_sheet(wb, "TableD_OBS_brute", td_obs)

    wsd2 = _write_df_to_sheet(wb, "TableD_NEMO", td_nem)
    _merge_cells_for_tableD(wsd2, sea_col=1, start_row=2)
    _write_df_to_sheet(wb, "TableD_NEMO_brute", td_nem)

    # =============================================================================
    # LAT×DECADE — 4 feuilles
    # =============================================================================
    wldo = _write_df_to_sheet(wb, "LatDec_OBS", latdec_obs)
    _merge_cells_for_latband_decade(wldo, band_col=1, start_row=2)
    _write_df_to_sheet(wb, "LatDec_OBS_brute", latdec_obs)

    wldn = _write_df_to_sheet(wb, "LatDec_NEMO", latdec_nem)
    _merge_cells_for_latband_decade(wldn, band_col=1, start_row=2)
    _write_df_to_sheet(wb, "LatDec_NEMO_brute", latdec_nem)

    # =============================================================================
    # LAT×PERIODE×PROFONDEUR — 4 feuilles
    # =============================================================================
    wlpdo = _write_df_to_sheet(wb, "LatPeriodDepth_OBS", latpd_obs)
    _merge_cells_for_latband_period_depth(wlpdo, band_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "LatPeriodDepth_OBS_brute", latpd_obs)

    wlpdn = _write_df_to_sheet(wb, "LatPeriodDepth_NEMO", latpd_nem)
    _merge_cells_for_latband_period_depth(wlpdn, band_col=1, period_col=2, start_row=2)
    _write_df_to_sheet(wb, "LatPeriodDepth_NEMO_brute", latpd_nem)

    # --- Scatter SO (propre + brute) + AJOUT shallow
    if enable_scatter_tables:
        # Scatter standard (inchangé)
        scatter_obs = _build_scatter_SO_table("OBS", obs_pts)
        scatter_nem = _build_scatter_SO_table("NEMO", nemo_pts)
        scatter_combined = pd.concat([scatter_obs, scatter_nem], ignore_index=True)
        scatter_combined = scatter_combined[['Dataset', 'Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]

        wso = _write_df_to_sheet(
            wb, "Scatter_SO_OBS",
            scatter_obs[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )
        _merge_cells_for_scatter_period(wso, period_col=1, start_row=2)
        _write_df_to_sheet(
            wb, "Scatter_SO_OBS_brute",
            scatter_obs[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )

        wsn = _write_df_to_sheet(
            wb, "Scatter_SO_NEMO",
            scatter_nem[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )
        _merge_cells_for_scatter_period(wsn, period_col=1, start_row=2)
        _write_df_to_sheet(
            wb, "Scatter_SO_NEMO_brute",
            scatter_nem[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )

        wsc = _write_df_to_sheet(wb, "Scatter_SO_Combined", scatter_combined)
        _merge_cells_for_scatter_combined(wsc, dataset_col=1, period_col=2, start_row=2)
        _write_df_to_sheet(wb, "Scatter_SO_Combined_brute", scatter_combined)

        # ----------------------------
        # AJOUT: Scatter shallow 10–50 et 50–100 (2 lignes par période)
        # ----------------------------
        shallow_depths = [(10, 50), (50, 100)]
        scatter_obs_sh = _build_scatter_SO_table("OBS", obs_pts, depths=shallow_depths)
        scatter_nem_sh = _build_scatter_SO_table("NEMO", nemo_pts, depths=shallow_depths)
        scatter_combined_sh = pd.concat([scatter_obs_sh, scatter_nem_sh], ignore_index=True)
        scatter_combined_sh = scatter_combined_sh[['Dataset', 'Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]

        # OBS shallow
        wso2 = _write_df_to_sheet(
            wb, "Scatter_SO_OBS_10-50_50-100",
            scatter_obs_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )
        _merge_cells_for_scatter_period(wso2, period_col=1, start_row=2)
        _write_df_to_sheet(
            wb, "Scatter_SO_OBS_10-50_50-100_brute",
            scatter_obs_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )

        # NEMO shallow
        wsn2 = _write_df_to_sheet(
            wb, "Scatter_SO_NEMO_10-50_50-100",
            scatter_nem_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )
        _merge_cells_for_scatter_period(wsn2, period_col=1, start_row=2)
        _write_df_to_sheet(
            wb, "Scatter_SO_NEMO_10-50_50-100_brute",
            scatter_nem_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]
        )

        # Combined shallow
        wsc2 = _write_df_to_sheet(wb, "Scatter_SO_Combined_10-50_50-100", scatter_combined_sh)
        _merge_cells_for_scatter_combined(wsc2, dataset_col=1, period_col=2, start_row=2)
        _write_df_to_sheet(wb, "Scatter_SO_Combined_10-50_50-100_brute", scatter_combined_sh)

    # =============================================================================
    # 3 CASE TABLES (OBS) — BA / Weddell / Ross
    # =============================================================================
    # 1) BA 1930–1939 : 10–200 + 130–160 avec z
    case_ba = _build_case_table_depth_years_with_z(
        obs_pts=obs_pts,
        sea_name='Bellingshausen-Amundsen',
        y0=1930,
        y1_inclusive=1939,
        depth_bands=[
            (10, 200, "10–200 m"),
            (130, 160, "130–160 m"),
        ],
    )

    # 2) Weddell 1937–1957 : 10–200 (simple)
    case_wed = _build_case_table_years_simple(
        obs_pts=obs_pts,
        sea_name='Weddell',
        y0=1937,
        y1_inclusive=1957,
        depth_band=(10, 200),
    )

    # 3) Ross 1920–1929 : 10–200 + 80–110 avec z (FIX z pooled => plus de colonne vide)
    case_ross = _build_case_table_depth_years_with_z(
        obs_pts=obs_pts,
        sea_name='Ross',
        y0=1920,
        y1_inclusive=1929,
        depth_bands=[
            (10, 200, "10–200 m"),
            (80, 110, "80–110 m"),
        ],
    )

    # --- Excel sheets
    if not case_ba.empty:
        ws_ba = _write_df_to_sheet(wb, "Case_BA_1930-1939", case_ba)
        _merge_cells_for_case_depth(ws_ba, depth_col=1, start_row=2)
        _write_df_to_sheet(wb, "Case_BA_1930-1939_brute", case_ba)

    if not case_wed.empty:
        _write_df_to_sheet(wb, "Case_Weddell_1937-1957", case_wed)
        _write_df_to_sheet(wb, "Case_Weddell_1937-1957_brute", case_wed)

    if not case_ross.empty:
        ws_ro = _write_df_to_sheet(wb, "Case_Ross_1920-1929", case_ross)
        _merge_cells_for_case_depth(ws_ro, depth_col=1, start_row=2)
        _write_df_to_sheet(wb, "Case_Ross_1920-1929_brute", case_ross)

    xlsx_name = f"StatsTables_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.xlsx"
    out_xlsx = out_dir / xlsx_name
    wb.save(out_xlsx)
    print(f"\n Excel stats généré : {out_xlsx.name}")

    # =============================================================================
    # PNG : Table1 + TableM + TableD (OBS/NEMO) + Scatter SO + LAT×DECADE + LAT×PERIODE×PROFONDEUR + CASES
    # =============================================================================
    if make_png:
        # ---- Table1
        out_png_t1_obs = out_dir / f"Table1_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        out_png_t1_nem = out_dir / f"Table1_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        _save_table1_like_png(t1_obs, out_png_t1_obs, font_size=10, decimals=PNG_DECIMALS)
        _save_table1_like_png(t1_nem, out_png_t1_nem, font_size=10, decimals=PNG_DECIMALS)
        print(f" PNG Table1 OBS : {out_png_t1_obs.name}")
        print(f" PNG Table1 NEMO: {out_png_t1_nem.name}")

        # ---- TableM
        out_png_tm_obs = out_dir / f"TableM_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        out_png_tm_nem = out_dir / f"TableM_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        _save_table1_like_png(tm_obs, out_png_tm_obs, font_size=13, decimals=PNG_DECIMALS)
        _save_table1_like_png(tm_nem, out_png_tm_nem, font_size=13, decimals=PNG_DECIMALS)
        print(f" PNG TableM OBS : {out_png_tm_obs.name}")
        print(f" PNG TableM NEMO: {out_png_tm_nem.name}")

        # ---- TableD
        out_png_td_obs = out_dir / f"TableD_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        out_png_td_nem = out_dir / f"TableD_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        _save_tableD_png(td_obs, out_png_td_obs, font_size=10, decimals=PNG_DECIMALS)
        _save_tableD_png(td_nem, out_png_td_nem, font_size=10, decimals=PNG_DECIMALS)
        print(f" PNG TableD OBS : {out_png_td_obs.name}")
        print(f" PNG TableD NEMO: {out_png_td_nem.name}")

        # ---- LAT×DECADE (OBS/NEMO)
        out_png_lat_obs = out_dir / f"LatDec_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        out_png_lat_nem = out_dir / f"LatDec_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        _save_table_latband_decade_png(latdec_obs, out_png_lat_obs, font_size=10, decimals=PNG_DECIMALS)
        _save_table_latband_decade_png(latdec_nem, out_png_lat_nem, font_size=10, decimals=PNG_DECIMALS)
        print(f" PNG LatDec OBS : {out_png_lat_obs.name}")
        print(f" PNG LatDec NEMO: {out_png_lat_nem.name}")

        # ---- LAT×PERIODE×PROFONDEUR (OBS/NEMO)
        out_png_latpd_obs = out_dir / f"LatPeriodDepth_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        out_png_latpd_nem = out_dir / f"LatPeriodDepth_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
        _save_table_latband_period_depth_png(latpd_obs, out_png_latpd_obs, font_size=10, decimals=PNG_DECIMALS)
        _save_table_latband_period_depth_png(latpd_nem, out_png_latpd_nem, font_size=10, decimals=PNG_DECIMALS)
        print(f" PNG LatPeriodDepth OBS : {out_png_latpd_obs.name}")
        print(f" PNG LatPeriodDepth NEMO: {out_png_latpd_nem.name}")

        # ---- Scatter SO (standard + shallow)
        if enable_scatter_tables:
            scatter_obs = _build_scatter_SO_table("OBS", obs_pts)
            scatter_nem = _build_scatter_SO_table("NEMO", nemo_pts)
            scatter_combined = pd.concat([scatter_obs, scatter_nem], ignore_index=True)
            scatter_combined = scatter_combined[['Dataset', 'Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]

            out_png_sc_obs = out_dir / f"ScatterSO_OBS_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            out_png_sc_nem = out_dir / f"ScatterSO_NEMO_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            out_png_sc_com = out_dir / f"ScatterSO_Combined_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"

            _save_table_png_generic(
                scatter_obs[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']],
                out_png_sc_obs,
                font_size=10,
                col_widths=[0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[{"col": "Période", "groupby": []}],
                decimals=PNG_DECIMALS
            )

            _save_table_png_generic(
                scatter_nem[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']],
                out_png_sc_nem,
                font_size=10,
                col_widths=[0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[{"col": "Période", "groupby": []}],
                decimals=PNG_DECIMALS
            )

            _save_table_png_generic(
                scatter_combined,
                out_png_sc_com,
                font_size=10,
                col_widths=[0.12, 0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[
                    {"col": "Dataset", "groupby": []},
                    {"col": "Période", "groupby": ["Dataset"]},
                ],
                decimals=PNG_DECIMALS
            )

            print(f" PNG Scatter SO OBS     : {out_png_sc_obs.name}")
            print(f" PNG Scatter SO NEMO    : {out_png_sc_nem.name}")
            print(f" PNG Scatter SO Combined: {out_png_sc_com.name}")

            # --- PNG shallow (10–50 / 50–100)
            shallow_depths = [(10, 50), (50, 100)]
            scatter_obs_sh = _build_scatter_SO_table("OBS", obs_pts, depths=shallow_depths)
            scatter_nem_sh = _build_scatter_SO_table("NEMO", nemo_pts, depths=shallow_depths)
            scatter_combined_sh = pd.concat([scatter_obs_sh, scatter_nem_sh], ignore_index=True)
            scatter_combined_sh = scatter_combined_sh[['Dataset', 'Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']]

            out_png_sc_obs_sh = out_dir / f"ScatterSO_OBS_10-50_50-100_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            out_png_sc_nem_sh = out_dir / f"ScatterSO_NEMO_10-50_50-100_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            out_png_sc_com_sh = out_dir / f"ScatterSO_Combined_10-50_50-100_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"

            _save_table_png_generic(
                scatter_obs_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']],
                out_png_sc_obs_sh,
                font_size=10,
                col_widths=[0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[{"col": "Période", "groupby": []}],
                decimals=PNG_DECIMALS
            )

            _save_table_png_generic(
                scatter_nem_sh[['Période', 'Profondeur', '⟨ΔT⟩', 'σ(ΔT)', '% ΔT < 0', '% ΔT > 0', 'N']],
                out_png_sc_nem_sh,
                font_size=10,
                col_widths=[0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[{"col": "Période", "groupby": []}],
                decimals=PNG_DECIMALS
            )

            _save_table_png_generic(
                scatter_combined_sh,
                out_png_sc_com_sh,
                font_size=10,
                col_widths=[0.12, 0.18, 0.14, 0.12, 0.12, 0.14, 0.14, 0.08],
                visual_merge=[
                    {"col": "Dataset", "groupby": []},
                    {"col": "Période", "groupby": ["Dataset"]},
                ],
                decimals=PNG_DECIMALS
            )

            print(f" PNG Scatter SO OBS (10–50/50–100)     : {out_png_sc_obs_sh.name}")
            print(f" PNG Scatter SO NEMO (10–50/50–100)    : {out_png_sc_nem_sh.name}")
            print(f" PNG Scatter SO Combined (10–50/50–100): {out_png_sc_com_sh.name}")

        # ---- PNG CASES
        if not case_ba.empty:
            out_png_case_ba = out_dir / f"Case_BA_1930-1939_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            _save_case_depth_table_png(case_ba, out_png_case_ba, font_size=12, decimals=PNG_DECIMALS)
            print(f" PNG Case BA : {out_png_case_ba.name}")

        if not case_wed.empty:
            out_png_case_wed = out_dir / f"Case_Weddell_1937-1957_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            _save_case_year_table_png(case_wed, out_png_case_wed, font_size=12, decimals=PNG_DECIMALS)
            print(f" PNG Case Weddell : {out_png_case_wed.name}")

        if not case_ross.empty:
            out_png_case_ross = out_dir / f"Case_Ross_1920-1929_{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}.png"
            _save_case_depth_table_png(case_ross, out_png_case_ross, font_size=12, decimals=PNG_DECIMALS)
            print(f" PNG Case Ross : {out_png_case_ross.name}")


# =============================================================================
# PANNEAU PRINCIPAL
# =============================================================================

def create_combined_panel(df_yamazaki, df_nemo, out_dir, season,
                          depth_min, depth_max, y0, y1,
                          enable_forensics=False,
                          enable_vertical_stats=False,
                          enable_new_tables=True,
                          tables_make_png=True,
                          enable_scatter_tables=True):
    """Crée le panneau combiné 4×2 avec logique 0-360°."""
    print(f"\n Génération du panneau combiné 4×2...")

    fig = plt.figure(figsize=(16, 18))
    gs = GridSpec(4, 2, figure=fig,
                  left=0.08, right=0.96,
                  top=0.97, bottom=0.06,
                  hspace=0.18, wspace=0.12)

    axes_yamazaki = []
    axes_nemo = []

    for i in range(4):
        axes_yamazaki.append(fig.add_subplot(gs[i, 0]))
        axes_nemo.append(fig.add_subplot(gs[i, 1]))

    season_months = SEASON_MONTHS[season]

    # Filtrage initial (base de TOUT ce qu’on sort en stats)
    df_yam = df_yamazaki[
        (df_yamazaki['hist_month'].isin(season_months)) &
        (df_yamazaki['hist_depth_m'] >= depth_min) &
        (df_yamazaki['hist_depth_m'] <= depth_max) &
        (df_yamazaki['hist_year'] >= y0) &
        (df_yamazaki['hist_year'] < y1)
    ].copy()

    df_nem = df_nemo[
        (df_nemo['hist_month'].isin(season_months)) &
        (df_nemo['hist_depth_m'] >= depth_min) &
        (df_nemo['hist_depth_m'] <= depth_max) &
        (df_nemo['hist_year'] >= y0) &
        (df_nemo['hist_year'] < y1)
    ].copy()

    print(f"   Yamazaki: {len(df_yam):,} observations")
    print(f"   NEMO: {len(df_nem):,} observations")

    # === TABLEAUX (Excel + PNG) ===
    if enable_new_tables:
        _generate_stats_outputs(
            df_yam_filtered=df_yam,
            df_nem_filtered=df_nem,
            out_dir=out_dir,
            season=season,
            depth_min=depth_min,
            depth_max=depth_max,
            y0=y0,
            y1=y1,
            make_png=tables_make_png,
            enable_scatter_tables=enable_scatter_tables
        )

    stats_rows = []  # vertical stats (optionnel)

    # Tracer chaque mer
    for i, sea_name in enumerate(SEA_ORDER):
        sea_config = SEAS[sea_name]

        df_sea_yam = filter_by_sea(df_yam, sea_config)
        df_sea_nem = filter_by_sea(df_nem, sea_config)

        print(f"\n {sea_name}:")
        print(f"   Yamazaki: {len(df_sea_yam):,} points")
        if len(df_sea_yam) > 0:
            print(f"      Lon range: {df_sea_yam['hist_lon'].min():.1f}° to "
                  f"{df_sea_yam['hist_lon'].max():.1f}° (format -180/180)")
            print(f"      Lon 0-360: {df_sea_yam['lon_360'].min():.1f}° to "
                  f"{df_sea_yam['lon_360'].max():.1f}°")

        print(f"   NEMO: {len(df_sea_nem):,} points")
        if len(df_sea_nem) > 0:
            print(f"      Lon range: {df_sea_nem['hist_lon'].min():.1f}° to "
                  f"{df_sea_nem['hist_lon'].max():.1f}° (format -180/180)")
            print(f"      Lon 0-360: {df_sea_nem['lon_360'].min():.1f}° to "
                  f"{df_sea_nem['lon_360'].max():.1f}°")

        _plot_yamazaki_sea_on_ax(axes_yamazaki[i], df_sea_yam, sea_name, season)
        _plot_nemo_sea_on_ax(axes_nemo[i], df_sea_nem, sea_name, season)

        sea_label = SEAS[sea_name]['label']
        axes_yamazaki[i].text(-0.12, 0.5, sea_label,
                              transform=axes_yamazaki[i].transAxes,
                              ha='right', va='center',
                              fontsize=14, fontweight='bold')

        axes_yamazaki[-1].set_xlabel('Anomalie de température (°C)', fontsize=13)
        axes_nemo[-1].set_xlabel('Anomalie de température (°C)', fontsize=13)

        axes_yamazaki[i].set_ylabel('Profondeur (m)', fontsize=13)

        if enable_vertical_stats:
            stats_row = _compute_vertical_stats_for_sea(
                sea_name,
                df_sea_yam,
                df_sea_nem,
                depth_min=depth_min,
                depth_max=depth_max
            )
            stats_rows.append(stats_row)

        if enable_forensics:
            _compute_profile_forensics_tables(
                sea_name=sea_name,
                df_sea_yam=df_sea_yam,
                df_sea_nem=df_sea_nem,
                out_dir=out_dir,
                depth_min=depth_min,
                depth_max=depth_max,
                bin_size=10,
                cell_res_deg=2,
                do_bootstrap=True,
                silent=False
            )

    fig.text(0.29, 0.985, 'Observations historiques',
             ha='center', va='top',
             fontsize=15, fontweight='bold')

    fig.text(0.74, 0.985, 'Simulation NEMO',
             ha='center', va='top',
             fontsize=15, fontweight='bold')

    legend_elements = [
        Line2D([0], [0], color=DECADE_COLORS[i], linewidth=2,
               marker='o', markersize=4,
               label=f'{dec0}-{dec1-1}')
        for i, (dec0, dec1) in enumerate(DECADES)
    ]

    fig.legend(handles=legend_elements,
               loc='lower center',
               bbox_to_anchor=(0.45, -0.01),
               ncol=7,
               fontsize=14,
               frameon=True,
               title='Décennies',
               title_fontsize=15)

    prefix = f"{season}_{depth_min}-{depth_max}m_{y0}-{y1-1}"
    out_path = out_dir / f"Combined_SeaAnomalies_{prefix}_4x2_panel.png"
    plt.savefig(out_path, dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()

    print(f"\n Panneau combiné sauvegardé : {out_path.name}")

    if enable_vertical_stats:
        _save_vertical_stats_table(stats_rows, out_dir, silent=False)


# =============================================================================
# MAIN
# =============================================================================

def main():
    ap = argparse.ArgumentParser(
        description="Panneau combiné 4x2 Yamazaki + NEMO (LON 0-360°)"
    )
    ap.add_argument("--yamazaki-csv-dir", type=str, required=True)
    ap.add_argument("--nemo-csv-dir", type=str, required=True)
    ap.add_argument("--out-dir", type=str, required=True)
    ap.add_argument("--season", type=str, default="DJF")
    ap.add_argument("--depth-min", type=int, default=10)
    ap.add_argument("--depth-max", type=int, default=200)
    ap.add_argument("--hist-y0", type=int, default=1900)
    ap.add_argument("--hist-y1", type=int, default=1970)

    # === FLAGS ===
    ap.add_argument("--enable-forensics", action="store_true",
                    help="Réactive les sorties 'forensics' (CSV + prints). OFF par défaut.")
    ap.add_argument("--enable-vertical-stats", action="store_true",
                    help="Réactive les vertical stats legacy (CSV + prints). OFF par défaut.")
    ap.add_argument("--disable-new-tables", action="store_true",
                    help="Désactive les tableaux Excel/PNG. ON par défaut.")
    ap.add_argument("--tables-no-png", action="store_true",
                    help="Ne génère pas les PNG (Excel uniquement).")
    ap.add_argument("--disable-scatter-tables", action="store_true",
                    help="Désactive les tableaux Scatter_SO (Excel + PNG). ON par défaut.")

    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 80)
    print("PANNEAU COMBINÉ 4×2 : Yamazaki + NEMO (LON 0-360°)")
    print("=" * 80)

    print("\n Chargement Yamazaki...")
    csv_pattern_yam = str(Path(args.yamazaki_csv_dir) /
                          "yamazaki_en4_wod_DJF_[0-9]*.csv")
    df_yamazaki = _safe_read_csvs(csv_pattern_yam)

    if df_yamazaki.empty:
        print(" Aucune donnée Yamazaki trouvée")
        return

    if 'yamazaki_T' not in df_yamazaki.columns:
        print(" Colonne 'yamazaki_T' manquante")
        return

    print(f"    Yamazaki chargé: {len(df_yamazaki):,} lignes")

    print("\n Chargement NEMO...")
    csv_pattern_nemo = str(Path(args.nemo_csv_dir) /
                           "nemo_yamazaki_DJF_[0-9]*.csv")
    df_nemo = _safe_read_csvs(csv_pattern_nemo)

    if df_nemo.empty:
        print(" Aucune donnée NEMO trouvée")
        return

    if ('nemo_recent_T' not in df_nemo.columns or
            'nemo_hist_T' not in df_nemo.columns):
        print(" Colonnes NEMO manquantes")
        return

    print(f"    NEMO chargé: {len(df_nemo):,} lignes")

    create_combined_panel(
        df_yamazaki, df_nemo, out_dir,
        args.season, args.depth_min, args.depth_max,
        args.hist_y0, args.hist_y1,
        enable_forensics=args.enable_forensics,
        enable_vertical_stats=args.enable_vertical_stats,
        enable_new_tables=(not args.disable_new_tables),
        tables_make_png=(not args.tables_no_png),
        enable_scatter_tables=(not args.disable_scatter_tables)
    )

    print(f"\n{'=' * 80}\n TERMINÉ\n{'=' * 80}")


if __name__ == "__main__":
    main()
